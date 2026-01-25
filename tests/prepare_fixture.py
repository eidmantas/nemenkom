"""
Helper script to prepare test fixture from real XLSX file
Uses existing project functions (fetcher, openpyxl)
"""
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))

from services.scraper.core.fetcher import fetch_xlsx, DEFAULT_URL
from openpyxl import load_workbook, Workbook

def prepare_fixture(num_rows=100):
    """Prepare test fixture from real XLSX file"""
    print('1. Fetching XLSX using existing fetcher...')
    temp_file = fetch_xlsx(DEFAULT_URL)
    
    # Read the actual Excel file
    wb = load_workbook(temp_file)
    ws = wb.active
    
    # Copy first N+1 rows (row 1 is empty, row 2 is header, rows 3+ are data)
    print(f'2. Copying first {num_rows + 1} rows (header + {num_rows} data rows)...')
    rows_to_copy = []
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=num_rows + 1, values_only=True), 1):
        rows_to_copy.append(row)
        if i == 2:
            print(f'   Header row: {row[:3]}...')
    
    # Create new workbook
    wb_out = Workbook()
    ws_out = wb_out.active
    
    # Write rows
    for row in rows_to_copy:
        ws_out.append(row)
    
    # Save fixture
    output_path = Path(__file__).parent / 'fixtures' / 'sample_schedule.xlsx'
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb_out.save(output_path)
    print(f'3. ✅ Created fixture: {output_path}')
    
    # Verify with parser
    import pandas as pd
    df_verify = pd.read_excel(output_path, skiprows=1)
    print(f'4. Verification: {len(df_verify)} rows')
    print(f'   Columns: {list(df_verify.columns)[:3]}...')
    print(f'   First Kaimai: {df_verify.iloc[0]["Kaimai"] if "Kaimai" in df_verify.columns else "NOT FOUND"}')
    
    temp_file.unlink()
    print('✅ Done!')
    return output_path

if __name__ == '__main__':
    prepare_fixture(100)
